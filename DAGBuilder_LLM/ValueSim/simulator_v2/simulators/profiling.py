from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from .common import TimingResult


@dataclass(frozen=True)
class ProfileEntry:
    key: str
    duration_s: float
    payload_bytes: float | None
    detail: str | None
    row: int


_UNIT_TO_SECONDS = {"s": 1.0, "ms": 1e-3, "us": 1e-6, "ns": 1e-9}


def _column_index(value: Any) -> int:
    if isinstance(value, int):
        if value <= 0:
            raise ValueError("Profiling column indexes are 1-based")
        return value
    from openpyxl.utils import column_index_from_string

    return column_index_from_string(str(value))


class ProfilingStore:
    def __init__(self, config: dict[str, Any], *, config_dir: Path | None = None) -> None:
        try:
            import openpyxl
        except ModuleNotFoundError as exc:
            raise ModuleNotFoundError(
                "openpyxl is required when simulator_v2 profiling is enabled"
            ) from exc

        path = Path(config["path"])
        if not path.is_absolute() and config_dir is not None:
            path = (config_dir / path).resolve()
        if not path.exists():
            raise FileNotFoundError(f"Profiling workbook does not exist: {path}")
        self.path = path
        self.sheet_name = str(config["sheet"])
        self.aliases = {str(key): str(value) for key, value in config.get("aliases", {}).items()}
        columns = config["columns"]
        key_col = _column_index(columns["key"])
        duration_col = _column_index(columns["duration"])
        payload_col = _column_index(columns["payload"]) if columns.get("payload") else None
        detail_col = _column_index(columns["detail"]) if columns.get("detail") else None
        scale = _UNIT_TO_SECONDS[str(config["duration_unit"])]
        first_row = int(config.get("header_rows", 1)) + 1

        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if self.sheet_name not in workbook.sheetnames:
            raise ValueError(f"Profiling workbook has no sheet named {self.sheet_name}")
        sheet = workbook[self.sheet_name]
        entries: dict[str, ProfileEntry] = {}
        duplicates: dict[str, list[int]] = {}
        for row in range(first_row, sheet.max_row + 1):
            raw_key = sheet.cell(row, key_col).value
            raw_duration = sheet.cell(row, duration_col).value
            if raw_key is None and raw_duration is None:
                continue
            if raw_key is None or raw_duration is None:
                raise ValueError(f"Profiling row {row} must contain both key and duration")
            key = str(raw_key).strip()
            payload_value = sheet.cell(row, payload_col).value if payload_col else None
            detail_value = sheet.cell(row, detail_col).value if detail_col else None
            entry = ProfileEntry(
                key=key,
                duration_s=float(raw_duration) * scale,
                payload_bytes=float(payload_value) if payload_value is not None else None,
                detail=str(detail_value) if detail_value is not None else None,
                row=row,
            )
            if key in entries:
                duplicates.setdefault(key, [entries[key].row]).append(row)
            else:
                entries[key] = entry
        if duplicates:
            rendered = ", ".join(f"{key}: rows {rows}" for key, rows in sorted(duplicates.items()))
            workbook.close()
            raise ValueError(f"Duplicate profiling keys: {rendered}")
        workbook.close()
        self.entries = entries

    def candidate_keys(self, node: dict[str, Any]) -> list[str]:
        values = [node.get("profile_key"), node.get("node_id"), node.get("op_name"), node.get("label")]
        keys: list[str] = []
        for raw in values:
            if raw is None:
                continue
            value = str(raw)
            mapped = self.aliases.get(value, value)
            if mapped not in keys:
                keys.append(mapped)
        return keys

    def find(self, node: dict[str, Any]) -> ProfileEntry | None:
        for key in self.candidate_keys(node):
            if key in self.entries:
                return self.entries[key]
        return None

    def require(self, node: dict[str, Any]) -> ProfileEntry:
        entry = self.find(node)
        if entry is None:
            raise KeyError(
                f"No profiling entry for {node.get('node_id')}; exact candidates={self.candidate_keys(node)}"
            )
        return entry

    def timing(self, node: dict[str, Any], category: str) -> TimingResult:
        entry = self.require(node)
        payload = entry.payload_bytes
        if payload is None and node.get("payload_bytes") is not None:
            payload = float(node["payload_bytes"])
        return TimingResult(
            duration_s=entry.duration_s,
            source="profiling",
            category=category,
            algorithm="exact_profile_key",
            payload_scope=str(node.get("payload_scope")) if node.get("payload_scope") else None,
            logical_payload_bytes=float(payload or 0.0),
            local_payload_bytes=float(payload or 0.0),
            detail={
                "profile_key": entry.key,
                "profile_xlsx": str(self.path),
                "profile_sheet": self.sheet_name,
                "profile_row": entry.row,
                "profile_detail": entry.detail,
            },
        )
